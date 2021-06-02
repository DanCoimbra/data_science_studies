import openpyxl as xl
from pprint import pprint

# Carrega as informações básicas do arquivo .xlsx
wb = xl.load_workbook('novoexcel.xlsx', data_only=True)
main = wb['Sheet1']
age = wb['Aging']
km = wb['km']
total_placa = wb.create_sheet('total_placa')
total_fipe_val = wb.create_sheet('total_fipe_val')
total_fipe_qtd = wb.create_sheet('total_fipe_qtd') 
total_fipe_avg = wb.create_sheet('total_fipe_avg')

def getval(sheet, coltext, row):
	for col in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
		header = str(sheet[f"{col}1"].value) 
		if header == coltext:
			value = sheet[f"{col}{row}"].value
			if value is not None:
				return str(value)

def cleanval(strval):
	strval = strval.replace("'", "")
	strval = strval.replace(".", "")
	index = 0
	while index < len(strval) and strval[index] != ',':
		index += 1
	strval = strval[:index]
	strval = int(strval)
	return strval

colmap = {
			"E" : "1",
			"F" : "2",
			"G" : "3",
			"H" : "4",
			"I" : "5",
			"J" : "6+",
			"K" : "10",
			"L" : "20",
			"M" : "30",
			"N" : "40",
			"O" : "50",
			"P" : "60",
			"Q" : "70",
			"R" : "80",
			"S" : "90",
			"T" : "100",
			"U" : "110+"
}

#################################### CAPTURA PLACAS ####################################
placas = dict()  # Dicionário cujas chaves são nomes de placa.
for row in range(2, main.max_row + 1):
	placa = getval(main, "Placa", row) # Nome da placa
	
	flag = 0
	if placa == "FYK5534":
		flag = 0

	if placa is not None:
		placas[placa] = placas.get(placa, dict()) # Cada placa tem 25 chaves

		for coltext in ["Fipe", "Montadora", "Série Simples"]: # Dados de identificação
			val = getval(main, coltext, row)
			if val is not None:
				placas[placa][coltext] = placas[placa].get(coltext, val)
			else:
				placas[placa][coltext] = placas[placa].get(coltext, "")

		placas[placa]["MPP"] = placas[placa].get("MPP", dict())
		placas[placa]["MP"] = placas[placa].get("MP", dict())
		placas[placa]["MC"] = placas[placa].get("MC", dict())
		placas[placa]["Total"] = placas[placa].get("Total", dict())

		for tipo in ["MPP", "MP", "MC", "Total"]:
			for coltext in ["1", "2", "3", "4", "5", "6+", "10", "20", "30", "40", "50", "60", "70", "80", "90", "100", "110+"]:
				placas[placa][tipo][coltext] = placas[placa][tipo].get(coltext, 0)

		tipo = getval(age, "Tipo", row)
		if tipo is not None:
			for coltext in ["1", "2", "3", "4", "5", "Vr Autorizado"]: # Dados de Aging
				val = getval(age, coltext, row)
				if val is not None:
					val = cleanval(val)
					if coltext != "Vr Autorizado":
						placas[placa][tipo][coltext] += val
						placas[placa]["Total"][coltext] += val
					else:
						placas[placa][tipo]["6+"] += val
						placas[placa]["Total"]["6+"] += val
					if (flag):
						print(val)
						print(coltext)
						pprint(placas[placa])
					break

			for coltext in ["10", "20", "30", "40", "50", "60", "70", "80", "90", "100", "Vr Autorizado"]:
				val = getval(km, coltext, row)
				if val is not None:
					val = cleanval(val)
					if coltext != "Vr Autorizado":
						placas[placa][tipo][coltext] += val
						placas[placa]["Total"][coltext] += val
					else:
						placas[placa][tipo]["110+"] += val
						placas[placa]["Total"]["110+"] += val
					if (flag):
						pprint(placas[placa])
					break

#################################### ESCREVE PLACAS ####################################
total_placa["A1"].value = "Fipe"
total_placa["B1"].value = "Montadora"
total_placa["C1"].value = "Série Simples"
total_placa["D1"].value = "Placa"
total_placa["E1"].value = "Tipo"
total_placa["F1"].value = "1 ano"
total_placa["G1"].value = "2 anos"
total_placa["H1"].value = "3 anos"
total_placa["I1"].value = "4 anos"
total_placa["J1"].value = "5 anos"
total_placa["K1"].value = "6+ anos"
total_placa["L1"].value = "10 kkm"
total_placa["M1"].value = "20 kkm"
total_placa["N1"].value = "30 kkm"
total_placa["O1"].value = "40 kkm"
total_placa["P1"].value = "50 kkm"
total_placa["Q1"].value = "60 kkm"
total_placa["R1"].value = "70 kkm"
total_placa["S1"].value = "80 kkm"
total_placa["T1"].value = "90 kkm"
total_placa["U1"].value = "100 kkm"
total_placa["V1"].value = "110+ kkm"

row = 2
for placa in placas:
	for newrow in range(row, row + 4):
		total_placa[f"A{newrow}"].value = placas[placa]["Fipe"]
		total_placa[f"B{newrow}"].value = placas[placa]["Montadora"]
		total_placa[f"C{newrow}"].value = placas[placa]["Série Simples"]
		total_placa[f"D{newrow}"].value = placa

	for index in range(0, 4):
		tipos = ["MPP", "MP", "MC", "Total"]
		total_placa[f"E{row + index}"].value = tipos[index]
		total_placa[f"F{row + index}"].value = placas[placa][tipos[index]]["1"]
		total_placa[f"G{row + index}"].value = placas[placa][tipos[index]]["2"]
		total_placa[f"H{row + index}"].value = placas[placa][tipos[index]]["3"]
		total_placa[f"I{row + index}"].value = placas[placa][tipos[index]]["4"]
		total_placa[f"J{row + index}"].value = placas[placa][tipos[index]]["5"]
		total_placa[f"K{row + index}"].value = placas[placa][tipos[index]]["6+"]
		total_placa[f"L{row + index}"].value = placas[placa][tipos[index]]["10"]
		total_placa[f"M{row + index}"].value = placas[placa][tipos[index]]["20"]
		total_placa[f"N{row + index}"].value = placas[placa][tipos[index]]["30"]
		total_placa[f"O{row + index}"].value = placas[placa][tipos[index]]["40"]
		total_placa[f"P{row + index}"].value = placas[placa][tipos[index]]["50"]
		total_placa[f"Q{row + index}"].value = placas[placa][tipos[index]]["60"]
		total_placa[f"R{row + index}"].value = placas[placa][tipos[index]]["70"]
		total_placa[f"S{row + index}"].value = placas[placa][tipos[index]]["80"]
		total_placa[f"T{row + index}"].value = placas[placa][tipos[index]]["90"]
		total_placa[f"U{row + index}"].value = placas[placa][tipos[index]]["100"]
		total_placa[f"V{row + index}"].value = placas[placa][tipos[index]]["110+"]

	row += 4

#################################### CAPTURA FIPES ####################################
fipes = dict()
for placa in placas:
	fipe = placas[placa]["Fipe"]
	fipes[fipe] = fipes.get(fipe, dict())
	
	fipes[fipe]["Montadora"] = fipes[fipe].get("Montadora", placas[placa]["Montadora"])
	fipes[fipe]["Série Simples"] = fipes[fipe].get("Série Simples", placas[placa]["Série Simples"])
	
	for tipo in ["MPP", "MP", "MC", "Total"]:
		fipes[fipe][tipo] = fipes[fipe].get(tipo, dict())
		for subtipo in ["1", "2", "3", "4", "5", "6+", "10", "20", "30", "40", "50", "60", "70", "80", "90", "100", "110+"]:
			fipes[fipe][tipo][subtipo] = fipes[fipe][tipo].get(subtipo, dict())
			fipes[fipe][tipo][subtipo]["VAL"] = fipes[fipe][tipo][subtipo].get("VAL", 0)
			fipes[fipe][tipo][subtipo]["QTD"] = fipes[fipe][tipo][subtipo].get("QTD", 0)
			valor = placas[placa][tipo][subtipo]
			if valor != 0:
				fipes[fipe][tipo][subtipo]["VAL"] += valor
				fipes[fipe][tipo][subtipo]["QTD"] += 1

#################################### ESCREVE FIPES ####################################
for planilha in [total_fipe_val, total_fipe_qtd, total_fipe_avg]:
	flag = 0
	if planilha == total_fipe_val:
		x = "VAL"
	elif planilha == total_fipe_qtd:
		x = "QTD"
	else:
		flag = 1

	termo = "Qtd"
	planilha["A1"].value = "Fipe"
	planilha["B1"].value = "Montadora"
	planilha["C1"].value = "Série Simples"
	planilha["D1"].value = "Tipo"
	planilha["E1"].value = "1 ano"
	planilha["F1"].value = "2 anos"
	planilha["G1"].value = "3 anos"
	planilha["H1"].value = "4 anos"
	planilha["I1"].value = "5 anos"
	planilha["J1"].value = "6+ anos"
	planilha["K1"].value = "10 kkm"
	planilha["L1"].value = "20 kkm"
	planilha["M1"].value = "30 kkm"
	planilha["N1"].value = "40 kkm"
	planilha["O1"].value = "50 kkm"
	planilha["P1"].value = "60 kkm"
	planilha["Q1"].value = "70 kkm"
	planilha["R1"].value = "80 kkm"
	planilha["S1"].value = "90 kkm"
	planilha["T1"].value = "100 kkm"
	planilha["U1"].value = "110+ kkm"

	row = 2
	for fipe in fipes:
		for newrow in range(row, row + 4):
			planilha[f"A{newrow}"].value = fipe
			planilha[f"B{newrow}"].value = fipes[fipe]["Montadora"]
			planilha[f"C{newrow}"].value = fipes[fipe]["Série Simples"]

		for index in range(0, 4):
			tipos = ["MPP", "MP", "MC", "Total"]
			planilha[f"D{row + index}"].value = tipos[index]
			for col in colmap:
				if not flag:
					planilha[f"{col}{row + index}"].value = fipes[fipe][tipos[index]][colmap[col]][x]
				else:
					if fipes[fipe][tipos[index]][colmap[col]]["QTD"] != 0:
						planilha[f"{col}{row + index}"].value = fipes[fipe][tipos[index]][colmap[col]]["VAL"] / fipes[fipe][tipos[index]][colmap[col]]["QTD"]
					else:
						planilha[f"{col}{row + index}"].value = 0

		row += 4

#################################### FINALIZA ####################################
del wb['Sheet1']
del wb['Aging']
del wb['km']

wb.save('novoteste.xlsx')