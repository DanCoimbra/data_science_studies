import openpyxl as xl
from pprint import pprint

# Carrega as informações básicas do arquivo .xlsx
wb = xl.load_workbook('novoexcel.xlsx', data_only=True)
main = wb['Sheet1']
age = wb['Aging']
km = wb['km']
total_placa = wb.create_sheet('total_placa')
total_série = wb.create_sheet('total_série')

def getval(sheet, coltext, row):
	for col in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
		if sheet[f"{col}1"].value == coltext:
			return sheet[f"{col}{row}"].value

# Captura as placas existentes
placas = dict()  # Dicionário cujas chaves são nomes de placa.
for row in range(2, start.max_row + 1):
	placa = getval(main, "Placa", row) # Nome da placa
	if placa is not None:
		placas[placa] = placas.get(placa, dict()) # Cada placa tem 25 chaves

		for coltext in ["Fipe", "Montadora", "Série Simples"]: # Dados de identificação
			val = getval(main, coltext, row)
			if val is not None:
				placas[placa][coltext] = placas[placa].get(coltext, val)
			else:
				placas[placa][coltext] = placas[placa].get(coltext, "")

		placas[placa]["MPP"] = placas[placa].get("MPP", 0)
		placas[placa]["MP"] = placas[placa].get("MP", 0)
		placas[placa]["MC"] = placas[placa].get("MPC", 0)
		placas[placa]["Total"] = placas[placa].get("Total", 0)

		tipo = getval(age, "Tipo", row)
		for coltext in ["1", "2", "3", "4", "5"]: # Dados de Aging
			val = getval(age, coltext, row)
			if val is not None:
				placas[placa][tipo] += val
				placas[placa]["Total"] += val
				break
		

		for col in "NOPQRSTUVWXY": # Dados numéritos
			lista_placas_cum[placa][col] = lista_placas_cum[placa].get(col, dict()) # Cada quilometragem tem 4 tipos de manutenção
			lista_placas_ind[placa][col] = lista_placas_ind[placa].get(col, dict())
			for tipo in ["MPP", "MP", "MC", "TOTAL"]:
				lista_placas_cum[placa][col][tipo] = lista_placas_cum[placa][col].get(tipo, 0);
				lista_placas_ind[placa][col][tipo] = lista_placas_ind[placa][col].get(tipo, 0);
			manutenção = start[f"J{row}"].value # Pega o tipo de manutenção
			valor = start[f"{col}{row}"].value # Pega a quilometragem
			if valor is not None: # Forma os sub-totais e o total de quilometragem por tipo de manutenção
				lista_placas_cum[placa][col]["TOTAL"] += int(valor)
				if manutenção == "Manutencao Preventiva Programada":
					lista_placas_cum[placa][col]["MPP"] += int(valor)
				elif manutenção == "Manutencao Preventiva":
					lista_placas_cum[placa][col]["MP"] += int(valor)
				elif manutenção == "Manutencao Corretiva":
					lista_placas_cum[placa][col]["MC"] += int(valor)

# Calcula os valores não-acumulados de cada placa.
for placa in lista_placas_cum:
	for col in "NOPQRSTUVWXY":
		for tipo in ["MPP", "MP", "MC", "TOTAL"]:
			lista_placas_ind[placa][col][tipo] = lista_placas_cum[placa][col][tipo] # Começa com o valor acumulado
			for prev_col in "NOPQRSTUVWXY".split(col)[0]:
				lista_placas_ind[placa][col][tipo] -= lista_placas_ind[placa][prev_col][tipo] # Subtrai os anteriores

# Inicializa as listas de séries
lista_séries_cum = dict()  # Valor cumulativo
lista_séries_ind = dict()  # Valor individual
for row in range(2, start.max_row + 1):
	fipe = start[f"A{row}"].value # Código FIPE
	montadora = start[f"B{row}"].value
	série = start[f"C{row}"].value # Código série
	if série is not None:
		placa = start[f"E{row}"].value # Captura placa à ser adicionada à lista de placas de cada série
		if série not in lista_séries_cum:
			lista_séries_cum[série] = dict()
			lista_séries_ind[série] = dict()
			lista_séries_cum[série]["A"] = lista_séries_cum[série].get("A", "000000-0")
			lista_séries_ind[série]["A"] = lista_séries_ind[série].get("A", "000000-0")
			x = int(fipe[:-2])
			y = int(lista_séries_cum[série]["A"][:-2])
			if x >= y:
				if x > y:
					lista_séries_cum[série]["A"] = fipe
					lista_séries_ind[série]["A"] = fipe
				if x == y:
					if int(fipe[-1]) > int(lista_séries_cum[série]["A"][-1]):
						ista_séries_cum[série]["A"] = fipe
						lista_séries_ind[série]["A"] = fipe
			lista_séries_cum[série]["B"] = montadora
			lista_séries_ind[série]["B"] = montadora
			for col in "NOPQRSTUVWXY": # Quilometragens de cada série
				lista_séries_cum[série][col] = dict()
				lista_séries_ind[série][col] = dict()
				for tipo in ["MPP", "MP", "MC", "TOTAL"]: # Quilometragens por tipo de manutenção de cada série
					lista_séries_cum[série][col][tipo] = lista_séries_cum[série][col].get(tipo, dict())
					lista_séries_ind[série][col][tipo] = lista_séries_ind[série][col].get(tipo, dict())
					lista_séries_cum[série][col][tipo]["val"] = 0
					lista_séries_cum[série][col][tipo]["qtd"] = 0
					lista_séries_ind[série][col][tipo]["val"] = 0
					lista_séries_ind[série][col][tipo]["qtd"] = 0
		if placa is not None:
			lista_séries_cum[série]["E"] = lista_séries_cum[série].get("E", list())
			if placa not in lista_séries_cum[série]["E"]:
				lista_séries_cum[série]["E"].append(placa) # Adiciona placas novas

# Condensa as quilometragens acumuladas por série, sem média.
for série in lista_séries_cum:
	for col in "NOPQRSTUVWXY": # Para cada quilometragem...
		for tipo in ["MPP", "MP", "MC", "TOTAL"]:
			for placa in lista_séries_cum[série]["E"]:
				lista_séries_cum[série][col][tipo]["val"] += lista_placas_cum[placa][col][tipo]
				if lista_placas_cum[placa][col][tipo] != 0:
					lista_séries_cum[série][col][tipo]["qtd"] += 1 # Para o cômputo da média, considera-se apenas carros com quilometragem não-nula
			
# Calcula os valores não-cumulados por série, sem média.
for série in lista_séries_cum:
	for col in "NOPQRSTUVWXY":
		for tipo in ["MPP", "MP", "MC", "TOTAL"]:
			lista_séries_ind[série][col][tipo]["val"] = lista_séries_cum[série][col][tipo]["val"] # Começa com o valor acumulado
			for prev_col in "NOPQRSTUVWXY".split(col)[0]:
				lista_séries_ind[série][col][tipo]["val"] -= lista_séries_ind[série][prev_col][tipo]["val"] # Subtrai os anteriores

# Calcula as quantidades de carros não-cumuladas por série.
for série in lista_séries_cum:
	for col in "NOPQRSTUVWXY": # Para cada quilometragem...
		for tipo in ["MPP", "MP", "MC", "TOTAL"]:
			for placa in lista_séries_cum[série]["E"]:
				if lista_placas_ind[placa][col][tipo] != 0:
					lista_séries_ind[série][col][tipo]["qtd"] += 1 # Para o cômputo da média, considera-se apenas carros com quilometragem não-nula
				
'''
# Calcula as médias dos séries.
for série in lista_séries_cum:
	for col in "NOPQRSTUVWXY":
		for tipo in ["MPP", "MP", "MC", "TOTAL"]:
			if lista_séries_cum[série][col][tipo]["qtd"] > 1:
				lista_séries_cum[série][col][tipo]["val"] = int(lista_séries_cum[série][col][tipo]["val"] / lista_séries_cum[série][col][tipo]["qtd"])
			if lista_séries_ind[série][col][tipo]["qtd"] > 1:
				lista_séries_ind[série][col][tipo]["val"] = int(lista_séries_ind[série][col][tipo]["val"] / lista_séries_ind[série][col][tipo]["qtd"])
'''

rel = {"N": "N",
		"O": "P",
		"P": "R",
		"Q": "T",
		"R": "V",
		"S": "X",
		"T": "Z",
		"U": "AB",
		"V": "AD",
		"W": "AF",
		"X": "AH",
		"Y": "AJ"
	}

# Copia o cabeçalho
for col in "ABCDEFGHIJKLM":
	cell = f"{col}1"
	valor = start[cell].value
	if valor is not None:
		total_placa[cell].value = valor
		total_série[cell].value = valor
for col in "NOPQRSTUVWXY":
	cell = f"{col}1"
	valor = start[cell].value
	if valor is not None:
		rel_col = rel[col]
		rel_cell = f"{rel_col}1"
		total_placa[cell].value = valor
		total_série[rel_cell].value = valor
for col in ["O", "Q", "S", "U", "W", "Y", "AA", "AC", "AE", "AG", "AI", "AK"]:
	cell = f"{col}1"
	total_série[cell].value = "Qtd"

# Em 'total_placa', escreve quatro linhas por placa: os valores não-acumulados
row = 2
for placa in lista_placas_ind:
	for col in "ABCDEFG":
		for i in range(0,4):
			cell = f"{col}{row + i}"
			value = lista_placas_cum[placa][col]
			if value == 0:
				total_placa[cell].value = "\"\""
			else:
				total_placa[cell].value = value

	prev_row = row
	for col in "NOPQRSTUVWXY":
		row = prev_row
		for tipo in ["MPP", "MP", "MC", "TOTAL"]:
			cell = f"{col}{row}"
			total_placa[f"J{row}"].value = tipo
			value = lista_placas_ind[placa][col][tipo]
			if value == 0:
				total_placa[cell].value = "\"\""
			else:
				total_placa[cell].value = value
			row += 1

# Em 'total_série', escreve quatro linhas por série: os valores não-acumulados e a quantia de carros
row = 2
for série in lista_séries_ind:
	prev_row = row
	for i in range(0, 4):
		total_série[f"A{row + i}"].value = lista_séries_ind[série]["A"] # Escreve o maior FIPE
		total_série[f"B{row + i}"].value = lista_séries_ind[série]["B"] # Escreve a montadora
		total_série[f"C{row + i}"].value = série # Escreve o nome do série

	for i in range(0, 12):
		row = prev_row
		org = ["N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y"]
		rel = ["N", "P", "R", "T", "V", "X", "Z", "AB", "AD", "AF", "AH", "AJ"]
		for tipo in ["MPP", "MP", "MC", "TOTAL"]:
			cell = f"{rel[i]}{row}"
			total_série[f"J{row}"].value = tipo
			value = lista_séries_ind[série][org[i]][tipo]["val"]
			if value == 0:
				total_série[cell].value = "\"\""
			else:
				total_série[cell].value = value
			row += 1

	for i in range(0, 12):
		row = prev_row
		org = ["N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y"]
		rel = ["O", "Q", "S", "U", "W", "Y", "AA", "AC", "AE", "AG", "AI", "AK"]
		for tipo in ["MPP", "MP", "MC", "TOTAL"]:
			next_cell = f"{rel[i]}{row}"
			value = lista_séries_ind[série][org[i]][tipo]["qtd"]
			if value == 0:
				total_série[next_cell].value = "\"\""
			else:
				total_série[next_cell].value = value
			row += 1

del wb['start']

wb.save('teste.xlsx')